# -*- coding: utf-8 -*-
"""Generate TEST_CASE_CURSOR sheet from checklist-ba.md spec."""
import openpyxl
from copy import copy

PATH = r"F:\5.May-2026\H2Q-SOLUTION-TESTER\Checklist Nhân sự\2. IT_TestCase-Template.xlsx"
SHEET_NAME = "TEST_CASE_CURSOR"

# Test cases: (section_header or None, id, description, procedure, expected, dependence, testdata)
# section_header rows only use description column
CASES = []

def add_section(name):
    CASES.append(("section", None, name, None, None, None, None))

def add_case(id_, desc, proc, expected, dep="", data=""):
    CASES.append(("case", id_, desc, proc, expected, dep, data))

# --- 1. Danh sách checklist ---
add_section("Danh sách checklist — /hrm/checklist")
add_case(
    "[CL-List-01]",
    "ADMIN truy cập màn danh sách checklist",
    "1. Đăng nhập tài khoản ADMIN\n2. Truy cập menu HRM → Checklist (route /hrm/checklist)",
    "1. Màn danh sách checklist hiển thị\n2. ADMIN thấy toàn bộ instance trong hệ thống\n3. Có đủ cột: Tên template, Nhân viên (fullName + accountName), StartDate, EndDate, Trạng thái, Tiến độ (done/total), Badge overdue (nếu có)",
    "1. Tài khoản ADMIN hợp lệ\n2. Đã có ít nhất 1 checklist instance",
)
add_case(
    "[CL-List-02]",
    "OFFICER chỉ thấy checklist liên quan mình",
    "1. Đăng nhập tài khoản OFFICER (có task được giao)\n2. Truy cập /hrm/checklist",
    "1. Danh sách chỉ hiển thị instance có ít nhất 1 task với AssignedOfficerId = OFFICER hiện tại\n2. Không hiển thị instance không liên quan",
    "1. OFFICER được giao task trong instance A\n2. Instance B không có task giao cho OFFICER này",
    "OFFICER: user có task assigned",
)
add_case(
    "[CL-List-03]",
    "EMPLOYEE chỉ thấy checklist của bản thân",
    "1. Đăng nhập tài khoản EMPLOYEE\n2. Truy cập /hrm/checklist",
    "1. Danh sách chỉ hiển thị instance có EmployeeId = user hiện tại\n2. Không hiển thị checklist của nhân viên khác",
    "1. EMPLOYEE có instance riêng\n2. Có instance của nhân viên khác trong DB",
)
add_case(
    "[CL-List-04]",
    "Hiển thị tiến độ doneTaskCount/totalTaskCount và % trên danh sách",
    "1. Truy cập /hrm/checklist với instance có 7/10 task done",
    "1. Cột tiến độ hiển thị 7/10 (hoặc 70%)\n2. Không hiển thị pendingTaskCount (đã bỏ v2)",
    "[CL-List-01]\nInstance: doneTaskCount=7, totalTaskCount=10",
)
add_case(
    "[CL-List-05]",
    "Hiển thị badge overdueTaskCount",
    "1. Truy cập danh sách với instance có task quá hạn chưa xử lý",
    "1. Instance hiển thị badge/số overdueTaskCount > 0\n2. Instance không overdue không hiển thị badge (hoặc = 0)",
    "Instance status Overdue hoặc có overdueTaskCount > 0",
)
add_case(
    "[CL-List-06]",
    "Filter danh sách theo Keyword",
    "1. Tại màn danh sách, nhập Keyword (tên nhân viên / tên template)\n2. Áp dụng filter",
    "1. Danh sách chỉ hiển thị record khớp Keyword\n2. API GET /api/checklist-assignment/getCheckList trả kết quả đã filter",
    "[CL-List-01]",
)
add_case(
    "[CL-List-07]",
    "Filter danh sách theo FromDate / ToDate",
    "1. Chọn FromDate và ToDate\n2. Áp dụng filter",
    "1. Chỉ hiển thị instance có StartDate/EndDate trong khoảng chọn\n2. Record ngoài khoảng không hiển thị",
    "[CL-List-01]",
)
add_case(
    "[CL-List-08]",
    "Filter danh sách theo Status",
    "1. Chọn Status (Pending/InProgress/Completed/Overdue/Cancelled)\n2. Áp dụng filter",
    "1. Danh sách chỉ hiển thị instance có status tương ứng",
    "[CL-List-01]\nCó instance với nhiều status khác nhau",
)
add_case(
    "[CL-List-09]",
    "Filter theo EmployeeId",
    "1. Chọn nhân viên cụ thể trong filter\n2. Áp dụng",
    "1. Chỉ hiển thị checklist của nhân viên được chọn",
    "[CL-List-01]",
)
add_case(
    "[CL-List-10]",
    "Phân trang danh sách checklist",
    "1. Truy cập danh sách khi có > số record trên 1 trang\n2. Chuyển trang 2, 3...",
    "1. Trang 1 hiển thị đúng số record theo page size\n2. Chuyển trang load đúng batch tiếp theo\n3. total count khớp API",
    "[CL-List-01]\nSố instance > page size",
)
add_case(
    "[CL-List-11]",
    "Click checklist → mở màn chi tiết",
    "1. Tại danh sách, click 1 dòng checklist",
    "1. Navigate tới /hrm/checklist/:instanceId\n2. Màn chi tiết load đúng instance",
    "[CL-List-01]",
)

# --- 2. Chi tiết checklist ---
add_section("Chi tiết checklist — /hrm/checklist/:instanceId")
add_case(
    "[CL-Detail-01]",
    "Hiển thị thông tin tổng quan instance",
    "1. Mở chi tiết 1 checklist instance",
    "1. Hiển thị: templateName, status, startDate, endDate\n2. endDate = MAX(deadline) các task\n3. progressPercent hiển thị (vòng tròn hoặc thanh %)",
    "[CL-List-11]",
)
add_case(
    "[CL-Detail-02]",
    "Hiển thị thông tin nhân viên (fullName, accountName, position, department)",
    "1. Mở chi tiết checklist",
    "1. Employee block hiển thị fullName từ user.FullName\n2. accountName = AccessControlProfileName ?? Username\n3. position và department hiển thị (JOIN bảng chức vụ/phòng ban)\n4. Không null như v1",
    "[CL-Detail-01]\nNhân viên có phòng ban và chức vụ trên hệ thống",
)
add_case(
    "[CL-Detail-03]",
    "Hiển thị sections và tasks theo SortOrder",
    "1. Mở chi tiết checklist có nhiều section/task",
    "1. Sections sắp theo sortOrder\n2. Tasks trong mỗi section sắp theo sortOrder\n3. Mỗi task: description, deadline, durationDays",
    "[CL-Detail-01]",
)
add_case(
    "[CL-Detail-04]",
    "Hiển thị officer resolved (fullName + accountName) mỗi task",
    "1. Mở chi tiết, xem task item",
    "1. Hiển thị officerFullName và officerAccountName (JOIN user)\n2. assignedOfficerId khớp task_item",
    "[CL-Detail-01]",
)
add_case(
    "[CL-Detail-05]",
    "Hiển thị trạng thái OfficerStatus / EmployeeStatus mỗi task",
    "1. Mở chi tiết với task ở các trạng thái khác nhau (Pending/Confirmed/Late/Undone)",
    "1. UI hiển thị đúng trạng thái officer và employee độc lập\n2. Hiển thị officerNote, employeeNote, lateReason nếu có\n3. Hiển thị confirmedAt tương ứng",
    "[CL-Detail-01]\nTask mix nhiều status",
)
add_case(
    "[CL-Detail-06]",
    "Hiển thị canOfficerUndo và officerUndoDeadline",
    "1. Mở task officer đã confirm < 60 phút, employee chưa confirm\n2. Mở task officer confirm > 60 phút hoặc employee đã confirm",
    "1. Case 1: hiển thị nút Undo + deadline undo\n2. Case 2: không hiển thị Undo (canOfficerUndo=false)",
    "[CL-Detail-01]",
)
add_case(
    "[CL-Detail-07]",
    "Hiển thị nhật ký (logs) read-only",
    "1. Mở chi tiết checklist đã có log (CreateInstance, Confirm, ChangeOfficer...)",
    "1. Danh sách log hiển thị createdAt, actorFullName, actorAccountName, action, detail\n2. Không cho sửa/xóa log",
    "[CL-Detail-01]\nInstance đã có nhiều log",
)
add_case(
    "[CL-Detail-08]",
    "Phân quyền xem chi tiết — OFFICER/EMPLOYEE không xem instance không liên quan",
    "1. OFFICER truy cập URL chi tiết instance không có task mình\n2. EMPLOYEE truy cập instance của người khác",
    "1. Không cho xem (403/redirect/empty)\n2. Không leak dữ liệu",
    "[CL-Detail-01]\nCó instance không liên quan",
)

# --- 3. Quản lý template ---
add_section("Quản lý template — /hrm/checklist/template (ADMIN)")
add_case(
    "[CL-Tpl-01]",
    "ADMIN truy cập màn quản lý template",
    "1. Đăng nhập ADMIN\n2. Truy cập /hrm/checklist/template",
    "1. Màn danh sách template hiển thị\n2. OFFICER/EMPLOYEE không truy cập được (hoặc không thấy menu)",
    "Tài khoản ADMIN",
)
add_case(
    "[CL-Tpl-02]",
    "Danh sách template — filter và phân trang",
    "1. Tại màn template, dùng filter và chuyển trang",
    "1. GET /api/checklist-template/getList trả đúng filter + pagination\n2. UI hiển thị Code, Name, IsActive",
    "[CL-Tpl-01]",
)
add_case(
    "[CL-Tpl-03]",
    "Tạo template mới — section + task + DefaultOfficerId",
    "1. Click Tạo template\n2. Nhập Code (không dấu), Name, Description\n3. Thêm section (Name, SortOrder)\n4. Thêm task (Description, DurationDays, SortOrder, DefaultOfficerId)\n5. Lưu",
    "1. POST /insert thành công\n2. Template lưu đủ section/task\n3. DefaultOfficerId lưu per task\n4. Code unique, không dấu",
    "[CL-Tpl-01]\nCode VD: BHS_NhanSuMoi_2025",
)
add_case(
    "[CL-Tpl-04]",
    "Validate Code template trùng / có dấu",
    "1. Tạo template Code đã tồn tại\n2. Tạo template Code có dấu/ký tự không hợp lệ",
    "1. Hiển thị lỗi validate, không tạo\n2. DB không duplicate Code",
    "[CL-Tpl-03]",
)
add_case(
    "[CL-Tpl-05]",
    "Xem chi tiết template",
    "1. Click 1 template trong danh sách",
    "1. GET /getDetail trả sections + tasks + DefaultOfficerId\n2. UI hiển thị đầy đủ cấu trúc template",
    "[CL-Tpl-01]",
)
add_case(
    "[CL-Tpl-06]",
    "Sửa template khi KHÔNG có instance active (status 0,1,3)",
    "1. Chọn template chưa có instance Pending/InProgress/Overdue\n2. Sửa Name/task/section → Lưu",
    "1. POST /update thành công\n2. Thay đổi được reflect",
    "[CL-Tpl-05]\nTemplate không có instance active",
)
add_case(
    "[CL-Tpl-07]",
    "Không cho sửa template khi CÓ instance active",
    "1. Chọn template đã có instance status 0/1/3\n2. Thử sửa và lưu",
    "1. BE từ chối (BR-08)\n2. UI hiển thị thông báo lỗi / disable edit",
    "[CL-Tpl-05]\nTemplate có instance InProgress",
)
add_case(
    "[CL-Tpl-08]",
    "Bật/tắt template (toggleActive)",
    "1. Toggle IsActive template từ ON → OFF\n2. Toggle OFF → ON",
    "1. PATCH /toggleActive cập nhật IsActive\n2. Template OFF không dùng được khi tạo instance mới (nếu FE enforce)",
    "[CL-Tpl-01]",
)

# --- 4. Tạo checklist instance ---
add_section("Tạo checklist instance (ADMIN)")
add_case(
    "[CL-Create-01]",
    "ADMIN tạo checklist — 1 nhân viên",
    "1. ADMIN chọn Tạo checklist\n2. Chọn Template + 1 Employee + StartDate\n3. (Nếu cần) chọn fallback AssignedOfficer\n4. Submit",
    "1. POST /insertBulk thành công\n2. Sinh checklist_instance + task_item\n3. Deadline mỗi task = StartDate + DurationDays\n4. EndDate instance = MAX deadline\n5. Status = Pending (0)",
    "[CL-Tpl-03]\nTemplate active, employee chưa có instance trùng",
    "templateId, startDate, employeeIds[1]",
)
add_case(
    "[CL-Create-02]",
    "ADMIN tạo checklist — nhiều nhân viên (bulk)",
    "1. Chọn cùng template + StartDate + nhiều employeeIds\n2. Submit",
    "1. Tạo N instance (1 per employee)\n2. Mỗi instance có đủ task từ template",
    "[CL-Create-01]\nemployeeIds: 2+ nhân viên",
)
add_case(
    "[CL-Create-03]",
    "Gán officer per-task từ DefaultOfficerId template",
    "1. Tạo instance từ template task có DefaultOfficerId\n2. Kiểm tra task_item.AssignedOfficerId",
    "1. AssignedOfficerId = template_task.DefaultOfficerId (BR-12)\n2. Không cần gán tay từng task",
    "[CL-Tpl-03]\nTask có DefaultOfficerId",
)
add_case(
    "[CL-Create-04]",
    "Fallback assignedOfficerId khi DefaultOfficerId = null",
    "1. Template có task DefaultOfficerId = null\n2. Tạo instance với assignedOfficerId fallback\n3. Kiểm tra task null officer",
    "1. AssignedOfficerId = request.assignedOfficerId cho task null DefaultOfficerId",
    "[CL-Tpl-03]\nTask không có DefaultOfficerId",
)
add_case(
    "[CL-Create-05]",
    "Validate fail — cả DefaultOfficerId và fallback đều null",
    "1. Template có task không DefaultOfficerId\n2. Tạo instance không truyền assignedOfficerId",
    "1. API validate fail\n2. Không tạo instance",
    "[CL-Tpl-03]",
)
add_case(
    "[CL-Create-06]",
    "Không tạo duplicate instance (BR-06)",
    "1. Employee đã có instance cùng TemplateId, Status ∈ {0,1,3}\n2. Tạo lại cùng template + employee",
    "1. API từ chối / báo lỗi duplicate\n2. Không tạo instance mới",
    "[CL-Create-01]\nInstance active đã tồn tại",
)
add_case(
    "[CL-Create-07]",
    "Notification khi tạo instance — gửi OFFICER được giao",
    "1. Tạo instance mới\n2. Kiểm tra in-app + email (SignalR)",
    "1. OFFICER được giao nhận: \"Bạn được giao checklist [Template] cho [Employee]\"",
    "[CL-Create-01]\nOFFICER có account active",
)

# --- 5. Officer confirm ---
add_section("Xác nhận task — Officer")
add_case(
    "[CL-Officer-01]",
    "OFFICER confirm đúng hạn (trước Deadline)",
    "1. Đăng nhập OFFICER\n2. Mở task chưa confirm, deadline chưa tới\n3. Nhập minh chứng (officerNote) nếu cần\n4. Confirm",
    "1. POST /officerConfirm thành công\n2. OfficerStatus = Confirmed (1)\n3. OfficerConfirmedAt/By được ghi\n4. Log OfficerConfirm\n5. Employee vẫn có thể confirm độc lập (BR-01)",
    "[CL-Create-01]\nOFFICER assigned, trước deadline",
)
add_case(
    "[CL-Officer-02]",
    "OFFICER confirm muộn — bắt buộc LateReason",
    "1. Mở task đã quá Deadline (hoặc confirm sau deadline)\n2. Confirm không nhập LateReason\n3. Confirm có LateReason",
    "1. Step 2: validate fail (BR-03)\n2. Step 3: OfficerStatus = Late (3), LateReason lưu, log OfficerConfirmLate",
    "[CL-Officer-01]\nTask quá deadline",
)
add_case(
    "[CL-Officer-03]",
    "OFFICER undo trong 60 phút — employee chưa confirm task",
    "1. Officer đã confirm task\n2. Trong 60 phút, EmployeeStatus task đó = 0\n3. Click Undo",
    "1. PATCH /officerUndo thành công\n2. OfficerStatus = Undone (2)\n3. Log OfficerUndo\n4. canOfficerUndo = true trước khi undo",
    "[CL-Officer-01]\nEmployee chưa confirm",
)
add_case(
    "[CL-Officer-04]",
    "OFFICER không undo sau 60 phút",
    "1. Officer confirm > 60 phút trước\n2. Employee chưa confirm\n3. Thử Undo",
    "1. Undo bị từ chối\n2. canOfficerUndo = false",
    "[CL-Officer-03]\nChờ > 60 phút hoặc mock time",
)
add_case(
    "[CL-Officer-05]",
    "OFFICER không undo khi employee đã confirm",
    "1. Officer confirm task\n2. Employee confirm cùng task\n3. Officer thử Undo",
    "1. Undo bị từ chối (BR-02)\n2. Task vẫn done",
    "[CL-Officer-01]\nEmployee confirm sau officer",
)
add_case(
    "[CL-Officer-06]",
    "EMPLOYEE confirm TRƯỚC OFFICER — xác nhận song song (BR-01)",
    "1. Task mới, officer chưa confirm\n2. EMPLOYEE confirm task\n3. OFFICER confirm cùng task",
    "1. Employee confirm thành công không cần officer confirm trước\n2. Sau cả 2 confirm → task done",
    "[CL-Create-01]",
)

# --- 6. Employee confirm ---
add_section("Xác nhận task — Employee")
add_case(
    "[CL-Employee-01]",
    "EMPLOYEE confirm đúng hạn",
    "1. Đăng nhập EMPLOYEE\n2. Mở task của instance bản thân, trước deadline\n3. Nhập employeeNote, Confirm",
    "1. POST /employeeConfirm thành công\n2. EmployeeStatus = Confirmed (1)\n3. Log EmployeeConfirm",
    "[CL-Create-01]\nEMPLOYEE = instance employee",
)
add_case(
    "[CL-Employee-02]",
    "EMPLOYEE confirm muộn — bắt buộc LateReason",
    "1. Confirm task quá deadline không LateReason\n2. Confirm với LateReason",
    "1. Step 1: fail validate\n2. Step 2: EmployeeStatus = Late (3)",
    "[CL-Employee-01]\nSau deadline",
)
add_case(
    "[CL-Employee-03]",
    "EMPLOYEE không confirm task instance của người khác",
    "1. EMPLOYEE thử confirm taskItem không thuộc instance mình",
    "1. API từ chối (BR-10)\n2. Không thay đổi status",
    "[CL-Employee-01]\nTaskItem khác employeeId",
)

# --- 7. Đổi cán bộ ---
add_section("Đổi cán bộ phụ trách task")
add_case(
    "[CL-ChangeOfficer-01]",
    "OFFICER đổi cán bộ — task OfficerStatus Pending (0)",
    "1. OFFICER mở task chưa confirm\n2. Chọn đổi cán bộ → chọn officer mới\n3. Lưu",
    "1. PATCH /changeOfficer thành công\n2. AssignedOfficerId cập nhật\n3. Log ChangeOfficer (OldId→NewId)\n4. Notify OFFICER mới",
    "[CL-Create-01]\nOfficerStatus = 0",
)
add_case(
    "[CL-ChangeOfficer-02]",
    "Đổi cán bộ task OfficerStatus Undone (2)",
    "1. Task officer đã undo (status 2)\n2. Đổi cán bộ",
    "1. Cho phép đổi (BR-04)\n2. AssignedOfficerId cập nhật",
    "[CL-Officer-03]",
)
add_case(
    "[CL-ChangeOfficer-03]",
    "Không đổi cán bộ task đã Confirmed/Late (1,3)",
    "1. Task OfficerStatus = 1 hoặc 3\n2. Thử đổi cán bộ",
    "1. API từ chối (BR-04)\n2. AssignedOfficerId không đổi",
    "[CL-Officer-01]",
)
add_case(
    "[CL-ChangeOfficer-04]",
    "ADMIN đổi cán bộ task chưa confirm",
    "1. ADMIN đổi AssignedOfficerId task status 0/2",
    "1. Thành công như OFFICER flow",
    "[CL-ChangeOfficer-01]\nTài khoản ADMIN",
)

# --- 8. Hủy checklist ---
add_section("Hủy checklist instance (ADMIN)")
add_case(
    "[CL-Cancel-01]",
    "ADMIN hủy instance — có CancelReason",
    "1. ADMIN mở instance\n2. Chọn Hủy, nhập lý do (nghỉ việc/chuyển bộ phận)\n3. Xác nhận",
    "1. PATCH /cancel thành công\n2. Status = Cancelled (4)\n3. CancelReason, CancelledBy/At lưu\n4. Log CancelInstance",
    "[CL-Create-01]\nADMIN account",
)
add_case(
    "[CL-Cancel-02]",
    "Hủy instance không có CancelReason",
    "1. ADMIN hủy không nhập lý do",
    "1. Validate fail (BR-07)\n2. Instance không chuyển Cancelled",
    "[CL-Cancel-01]",
)
add_case(
    "[CL-Cancel-03]",
    "Notification khi hủy — OFFICER + EMPLOYEE",
    "1. Hủy instance thành công",
    "1. OFFICER liên quan + EMPLOYEE nhận: \"Checklist [Template]/[Employee] đã hủy. Lý do: [CancelReason]\"",
    "[CL-Cancel-01]",
)
add_case(
    "[CL-Cancel-04]",
    "OFFICER/EMPLOYEE không được hủy instance",
    "1. OFFICER hoặc EMPLOYEE thử gọi cancel",
    "1. Từ chối — chỉ ADMIN (BR-07)",
    "[CL-Create-01]",
)

# --- 9. Tiến độ & trạng thái ---
add_section("Tiến độ và trạng thái instance")
add_case(
    "[CL-Status-01]",
    "Instance Pending (0) khi mới tạo — chưa ai confirm",
    "1. Tạo instance mới\n2. Kiểm tra status",
    "1. Status = Pending (0)\n2. progressPercent = 0",
    "[CL-Create-01]",
)
add_case(
    "[CL-Status-02]",
    "Instance InProgress (1) khi có bất kỳ confirm (officer HOẶC employee)",
    "1. Tạo instance\n2. Chỉ OFFICER confirm 1 task HOẶC chỉ EMPLOYEE confirm 1 task",
    "1. Status chuyển InProgress (1) — không cần cả 2 (v2)",
    "[CL-Create-01]",
)
add_case(
    "[CL-Status-03]",
    "Task done khi Officer ∈ {1,3} AND Employee ∈ {1,3}",
    "1. Confirm cả officer và employee cùng task (on-time hoặc late)",
    "1. Task được tính done\n2. doneTaskCount tăng, progressPercent cập nhật",
    "[CL-Officer-01]\n[CL-Employee-01]",
)
add_case(
    "[CL-Status-04]",
    "Instance Completed (2) khi tất cả task done",
    "1. Hoàn thành confirm tất cả task",
    "1. Status = Completed (2)\n2. Log InstanceCompleted\n3. Notify ADMIN + EMPLOYEE",
    "[CL-Status-03]\nTất cả task done",
)
add_case(
    "[CL-Status-05]",
    "Instance Overdue (3) — job ChecklistOverdueJob",
    "1. Có task Deadline < TODAY, OfficerStatus=0 hoặc EmployeeStatus=0\n2. Chạy job 00:05 (hoặc trigger manual)",
    "1. Bên chưa xử lý set Late (3) độc lập\n2. Instance Status = Overdue nếu đang 0/1\n3. Log TaskOverdue + notification",
    "[CL-Create-01]\nTask quá hạn, job enabled",
)
add_case(
    "[CL-Status-06]",
    "progressPercent = doneTaskCount / totalTaskCount * 100",
    "1. Instance 3/10 task done",
    "1. progressPercent = 30\n2. List và detail đồng bộ",
    "[CL-Status-03]",
)

# --- 10. Notification ---
add_section("Thông báo (Notification)")
add_case(
    "[CL-Notify-01]",
    "Nhắc task sắp đến hạn T-2",
    "1. Task deadline trong 2 ngày, bên chưa confirm",
    "1. Bên chưa confirm nhận: \"Task [TaskDesc] đến hạn [Deadline]\"",
    "Job/scheduler T-2 enabled",
)
add_case(
    "[CL-Notify-02]",
    "Thông báo task quá hạn",
    "1. Task quá deadline chưa xử lý",
    "1. Bên chưa confirm + ADMIN nhận overdue notification",
    "[CL-Status-05]",
)
add_case(
    "[CL-Notify-03]",
    "Không notify \"đến lượt employee\" (v2 bỏ thứ tự)",
    "1. Officer confirm task, employee chưa confirm",
    "1. Không có notification kiểu \"đến lượt employee confirm\"",
    "[CL-Officer-01]",
)

def copy_row_style(ws_src, ws_dst, src_row, dst_row, max_col=11):
    for c in range(1, max_col + 1):
        src = ws_src.cell(src_row, c)
        dst = ws_dst.cell(dst_row, c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = copy(src.number_format)

def main():
    wb = openpyxl.load_workbook(PATH)
    login = wb["Login"]

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    # Copy header rows 1-8 from Login
    for r in range(1, 9):
        for c in range(1, 12):
            val = login.cell(r, c).value
            ws.cell(r, c, val)
        copy_row_style(login, ws, r, r)
        if login.row_dimensions[r].height:
            ws.row_dimensions[r].height = login.row_dimensions[r].height

    # Copy merged cells in header
    for merge in login.merged_cells.ranges:
        if merge.min_row <= 8:
            ws.merge_cells(str(merge))

    # Module metadata rows 2-6
    ws.cell(2, 2, "Checklist Nhân sự")
    ws.cell(3, 2, "Verify module Checklist Nhân sự (template, instance, confirm song song, tiến độ v2)")
    ws.cell(4, 2, "hiennm23")

    case_count = sum(1 for x in CASES if x[0] == "case")
    ws.cell(6, 5, case_count)
    ws.cell(6, 1, 0)
    ws.cell(6, 2, 0)
    ws.cell(6, 3, case_count)
    ws.cell(6, 4, 0)

    row = 9
    for item in CASES:
        kind = item[0]
        if kind == "section":
            ws.cell(row, 2, item[2])
            copy_row_style(login, ws, 9, row)
            if login.row_dimensions[9].height:
                ws.row_dimensions[row].height = login.row_dimensions[9].height
            row += 1
        else:
            _, id_, desc, proc, expected, dep, data = item
            ws.cell(row, 1, id_)
            ws.cell(row, 2, desc)
            ws.cell(row, 3, proc)
            ws.cell(row, 4, expected)
            ws.cell(row, 5, dep)
            ws.cell(row, 6, data)
            ws.cell(row, 7, "Untested")
            copy_row_style(login, ws, 10, row)
            ws.row_dimensions[row].height = 80
            row += 1

    out = PATH
    try:
        wb.save(out)
    except PermissionError:
        out = r"C:\Users\Admin\checklist_TEST_CASE_CURSOR.xlsx"
        wb.save(out)
        print("Original file locked - saved copy:", out)
    print("Saved", case_count, "test cases to sheet", SHEET_NAME)

if __name__ == "__main__":
    main()
