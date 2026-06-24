# -*- coding: utf-8 -*-
"""[LEGACY] Export TEST_CASE_CURSOR sheet to data/testcases.json.

Không còn được gọi từ API/dashboard — chỉ dùng thủ công nếu cần tham chiếu Excel cũ.
Nguồn chính thức: scripts/generate_testcases_v2.py → data/testcases.json
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from api.paths import EXCEL_PATH
from api.services.testcase_store import save_testcases

SHEET = "TEST_CASE_CURSOR"


def export_testcases(excel_path: Path | None = None) -> dict:
    excel_path = excel_path or EXCEL_PATH

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[SHEET]

    section = ""
    rows: list[dict] = []
    for r in range(9, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        desc = ws.cell(r, 2).value
        if not cid and desc:
            section = str(desc)
            continue
        if cid and str(cid).startswith("["):
            rows.append(
                {
                    "id": str(cid).strip("[]"),
                    "section": section,
                    "description": str(desc or ""),
                    "procedure": str(ws.cell(r, 3).value or ""),
                    "expected": str(ws.cell(r, 4).value or ""),
                    "dependence": str(ws.cell(r, 5).value or ""),
                    "testData": str(ws.cell(r, 6).value or ""),
                    "result": str(ws.cell(r, 7).value or "Untested"),
                    "testDate": str(ws.cell(r, 8).value or ""),
                    "note": str(ws.cell(r, 9).value or ""),
                    "evidence": str(ws.cell(r, 10).value or ""),
                }
            )

    data = {
        "moduleCode": str(ws.cell(2, 2).value or ""),
        "requirement": str(ws.cell(3, 2).value or ""),
        "tester": str(ws.cell(4, 2).value or ""),
        "cases": rows,
    }
    return save_testcases(data)


if __name__ == "__main__":
    result = export_testcases()
    print(f"Exported {len(result['cases'])} cases -> data/testcases.json")
