# -*- coding: utf-8 -*-
"""Ghi kết quả Playwright từ log-playwright vào Excel TEST_CASE_CURSOR."""
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
EXCEL = ROOT / "2. IT_TestCase-Checklist-Cursor.xlsx"
SHEET = "TEST_CASE_CURSOR"
RESULTS_JSON = ROOT / "log-playwright" / "latest" / "results.json"
RUN_LOG_HINT = ROOT / "log-playwright" / "last_run.txt"


def load_results():
    path = RESULTS_JSON
    if RUN_LOG_HINT.exists():
        run_dir = Path(RUN_LOG_HINT.read_text(encoding="utf-8").strip())
        candidate = run_dir / "results.json"
        if candidate.exists():
            path = candidate
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    by_case = {}
    for _key, v in raw.items():
        cid = v["case_id"]
        by_case.setdefault(cid, []).append(v)
    return by_case, path


def detect_role(description: str) -> str | None:
    if not description:
        return None
    d = description.upper()
    if "ADMIN" in d and "OFFICER" not in d and "EMPLOYEE" not in d:
        return "ADMIN"
    if "OFFICER" in d or "CÁN BỘ" in d.upper() or "TRƯỞNG PHÒNG" in d.upper():
        return "OFFICER"
    if "EMPLOYEE" in d or "NHÂN VIÊN" in d.upper():
        return "EMPLOYEE"
    return None


def pick_result(case_id: str, description: str, entries: list) -> tuple[str, str, str]:
    """Return (result, note, test_date)."""
    if not entries:
        return "Untested", "", ""

    role = detect_role(description)
    if role:
        matched = [e for e in entries if e["role"] == role]
        if matched:
            entries = matched

    if len(entries) == 1:
        e = entries[0]
        return e["result"], e["note"], e["time"][:10]

    # Nhiều role — tổng hợp
    parts = [f"{e['role']}:{e['result']}" for e in entries]
    note = "; ".join(parts)
    if len(parts) > 1:
        detail = " | ".join(f"{e['role']}: {e['note'][:80]}" for e in entries)
        note = f"{note} — {detail}"[:500]

    if all(e["result"] == "Pass" for e in entries):
        result = "Pass"
    elif any(e["result"] == "Fail" for e in entries):
        result = "Fail"
    else:
        result = entries[0]["result"]

    test_date = entries[0]["time"][:10]
    return result, note[:500], test_date


def main():
    by_case, src = load_results()
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]

    stats = {"Pass": 0, "Fail": 0, "Untested": 0, "N/A": 0}
    updated = 0

    for row in range(9, ws.max_row + 1):
        cell_id = ws.cell(row, 1).value
        if not cell_id or not str(cell_id).startswith("[CL-"):
            continue

        case_id = str(cell_id).strip("[]")
        desc = ws.cell(row, 2).value or ""
        entries = by_case.get(case_id, [])
        result, note, test_date = pick_result(case_id, desc, entries)

        ws.cell(row, 7, result)
        if test_date:
            ws.cell(row, 8, test_date)
        if note:
            ws.cell(row, 9, note)

        stats[result] = stats.get(result, 0) + 1
        if entries:
            updated += 1

    total = sum(stats.values())
    ws.cell(6, 1, stats.get("Pass", 0))
    ws.cell(6, 2, stats.get("Fail", 0))
    ws.cell(6, 3, stats.get("Untested", 0))
    ws.cell(6, 4, stats.get("N/A", 0))
    ws.cell(6, 5, total)

    try:
        wb.save(EXCEL)
        out = EXCEL
    except PermissionError:
        out = EXCEL.with_name(EXCEL.stem + "_updated.xlsx")
        wb.save(out)

    print(f"Source: {src}")
    print(f"Updated {updated} cases with automation results")
    print(f"Pass={stats.get('Pass',0)} Fail={stats.get('Fail',0)} Untested={stats.get('Untested',0)}")
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
